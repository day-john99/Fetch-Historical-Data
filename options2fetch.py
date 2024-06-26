
import shutil

import os
import math

import time
import pandas as pd
import openpyxl
from datetime import datetime
from datetime import timedelta

from breeze_connect import BreezeConnect





class Pull2 :


    def __init__(self):




        #----------------------------------------------------------------
        self.expiry = "2021-11-25T07:00:00.000Z"                         # set date

        # setting start date
        expiry_date1 = datetime.fromisoformat(self.expiry[:-1])  # convert expiry ,from iso format to datetime object

        three_months = timedelta(
            days=30 * 2)  # Calculate no of days in two months (assuming 30 days per month for simplicity)

        # Subtract three months from the expiry_date1
        self.str_date = expiry_date1 - three_months  # this will be our start date in datetime obj format

        # convert start_date into iso format
        self.str_date = self.str_date.isoformat()

        self.str_date = self.str_date + '.000Z'       # to get perfect iso format

        # --------------------------------------------------------------------------
        # setting end date


        two_days = timedelta(days=2)

        temp = datetime.fromisoformat(self.str_date[:-1])
        # counter updates end date ,every 2 days
        self.end_date = temp  + two_days  # here end_date is datetime object

        # converting it into iso format
        self.end_date = self.end_date.isoformat()

        self.end_date = self.end_date + '.000Z'  # to get perfect iso format


    # ---------------------------------------------------------------------------

    def str_date_upgrade(self):

        self.temp_date = datetime.fromisoformat(self.end_date[:-1])  # converting iso into datetime obj

        one_day = timedelta(days=1)
        self.temp_date = self.temp_date + one_day

        # convert back to iso format
        self.temp_date = self.temp_date.isoformat()

        self.temp_date = self.temp_date + '.000Z'  # to get perfect iso format

        return  self.temp_date

    # ---------------------------------------------------------------------------

    def end_date_upgrade(self):

        self.end_date = datetime.fromisoformat(self.end_date[:-1])    # converting iso into datetime obj

        two_days = timedelta(days=2)

        # counter updates end date ,every 2 days
        self.end_date = self.end_date + two_days  # here start_date is datetime object

        # converting it into iso format
        self.end_date = self.end_date.isoformat()

        self.end_date = self.end_date + '.000Z'  # to get perfect iso format


     # condition to correct enddate as if enddate greater than expiry date by 1,2 day as its updated by 2 days always
        if( ( datetime.fromisoformat(self.end_date[:-1]) > datetime.fromisoformat(self.expiry[:-1]) )   ) :
              self.end_date = self.expiry

        return self.end_date

    # ---------------------------------------------------------------------------

# get logged in

# Initialize SDK
breeze = BreezeConnect(api_key="BREEZE_API_KEY")

# Generate Session
breeze.generate_session(api_secret="BREEZE_API_SECRET",
                                session_token="SESSION_TOKEN")

print(breeze.get_funds())
# successfully logged in

# -----------------------------------------------------------------------------------



obj = Pull2()                  # constructor initializes from_date and to_date
count = 0
file_no = 0
name = "CNXBAN"
c_p = "put"


#  -------------------
# options multiple strikes calculation
# input , month's opening price
month_open = 35000
month_open = int(month_open)  # converting from string to int type

strike_first = 39200  # for 10% minus off, month open price
strike_last = 40500  # for 10% plus off, month open price

strike_first = math.floor(strike_first / 100) * 100  # to round off values to hundreds
strike_last =  math.floor(strike_last / 100) * 100


# -------------------
day_limit = 0

df_full = pd.DataFrame()     # create an empty dataframe

print("\n",obj.str_date,"\n",obj.end_date)


while( (datetime.fromisoformat(obj.end_date)) <= (datetime.fromisoformat(obj.expiry)) ):

  timer_start = time.time()          # start timer
  count = count + 1

  try :
    df = breeze.get_historical_data(interval="1minute",
                                  from_date=obj.str_date,
                                  to_date=obj.end_date,
                                  stock_code=name,
                                  exchange_code="NFO",
                                  product_type="options",
                                  expiry_date=obj.expiry,
                                  right=c_p,
                                  strike_price=strike_first)

    df = pd.DataFrame(df["Success"])                     #piece of data

    df_full = df_full._append(df , ignore_index=True)    # appended complete data as CD


    day_limit = day_limit + 1                           # to check api day limit of, 5000 calls per day

  except Exception as e:
      print(f"Error encountered: {e}")
      time.sleep(5)  # Wait 10 seconds before retrying
      continue  # Continue to the next iteration of the loop


  if (count == 99  and  (time.time() - timer_start) <= 60)  :

       sleepy = 60 - (time.time() - timer_start)  +3           # here 3 sec act as buffer
       time.sleep(sleepy)                                      # sleep for sleepy second
       count = 0


  if ( (datetime.fromisoformat(obj.end_date)) == (datetime.fromisoformat(obj.expiry)) ) :  # conditn to end file one

      # Export the ,appended complete data or CD to an Excel file
      df_full.to_excel('datatemp.xlsx', index=False)  # excel file

      print("\n ended file  {}".format(file_no))
      print("\ncalled times: ",day_limit)

      print(pd.read_excel('datatemp.xlsx'))                # prints values in terminal to cross check


      #----------------------------------------------------------------------------------------data cleaning starts

      # Step 1: Read the Excel file
      file_path = 'datatemp.xlsx'  # Replace with your file path
      df = pd.read_excel(file_path)

      # Step 2: Clear data in specific columns from row 3 onward
      # Specify the columns from which you want to delete data
      columns_to_clear = ['stock_code', 'exchange_code', 'product_type', 'expiry_date', 'right',
                          'strike_price']  # Replace with your column names

      # Keep the first two rows intact and clear the rest in the specified columns
      df.loc[1:, columns_to_clear] = None  # You can use None or np.nan

      # Step 3: Write the modified DataFrame back to an Excel file
      # output_file_path = 'modified_excel_file.xlsx'  # Replace with your desired file path
      df.to_excel(file_path, index=False)  # Set index=False to avoid writing row indices to the Excel file




      # ---------------------------------------------------------------------------------------data cleaning ends


      # Load the source workbook thats abt to be copied
      source_workbook = openpyxl.load_workbook('datatemp.xlsx')

      # Get the active worksheet of the source workbook
      source_sheet = source_workbook.active

      # Create a new workbook
      new_workbook = openpyxl.Workbook()

      # Get the active worksheet of the new workbook
      new_sheet = new_workbook.active

      # Iterate over rows in the source worksheet and copy data to the new worksheet


      for row in source_sheet.iter_rows(values_only=True):
          new_sheet.append(row)

      #shutil.copy('datatemp.xlsx', '{}.xlsx'.format(file_no))

      # Save the new workbook to a new Excel file under name of,
      #new_workbook.save(f'{name}_{obj.strike_first}_{c_p}_{datetime.fromisoformat(obj.expiry)}.xlsx')       # main file
                                                                                   # to be saved after each one contract
      new_workbook.save('{}.xlsx'.format(file_no))




      #-------- data to be used for rename of file
      timestamp = obj.expiry

      # Parse the timestamp string into a datetime object
      dt_obj = datetime.strptime(timestamp, "%Y-%m-%dT%H:%M:%S.%fZ")

      # Format the datetime object into the desired string format
      formatted_date = dt_obj.strftime("%d %B %Y")

      #--------

      # rename file
      os.rename( '{}.xlsx'.format(file_no) , '{} {} {} {}.xlsx'.format(name,formatted_date,strike_first,c_p) )



      #------------------------------------------
      # now deleting the old datatemp.xlsx file
      if os.path.exists('datatemp.xlsx'):
          os.remove('datatemp.xlsx')

      # emptying old filled dataframes
      df = pd.DataFrame()  # empty old, filled dataframe
      df_full = pd.DataFrame()  # empty old, filled dataframe

      file_no = file_no + 1

      #      STRIKE UPDATIION

      strike_first = strike_first + 100
      print("\nstrike updated", strike_first)

      #          STRIKE UPDATION BLOC END

      # from date and to date are then reset
      obj2 = Pull2()        # its called to reset from_date and to_date once again,coz they remain same ,whatever the expiry

      obj.str_date = obj2.str_date
      obj.end_date = obj2.end_date


      #print("\nout of ct ,from date",obj.str_date,"\n",obj.end_date)

      if (strike_first > strike_last):                         # TO CHECK IF 320 SCRIPS ARE FETCHED THEN END
          print("----------------END OF CODE--------------")
          break



  # updating from_date and to_date for loop

  obj.str_date = obj.str_date_upgrade()         # new, start date for new iteration

  obj.end_date = obj.end_date_upgrade()         # new, end date for new iteration

# close
